from fastapi import FastAPI, Depends, HTTPException, Query
from fastapi.security import OAuth2PasswordRequestForm
from fastapi.responses import StreamingResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from sqlalchemy import or_, and_
from typing import List, Optional
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from io import BytesIO
import json
import traceback

from database import SessionLocal, engine
from models import Base, Documento, VersaoDocumento, EstadoDocumento, Utilizador, PerfilUtilizador, Notificacao, ComentarioDocumento
from schemas import (
    DocumentoCreate, DocumentoUpdate, DocumentoOut,
    VersaoOut, MudancaEstado, UtilizadorCreate, Token,
    PasswordUpdate, ComentarioCreate, ComentarioOut
)
from auth import (
    hash_password,
    verificar_password,
    criar_token_acesso,
    get_current_user,
    ACCESS_TOKEN_EXPIRE_MINUTES
)
from templates import PROCESSOS_PADRAO, get_processos_from_data, criar_estrutura_com_processos
from dashboard import (
    get_dashboard_kpis,
    get_top_parceiros,
    get_documentos_recentes
)
from notificacoes import (
    criar_notificacao_para_empresa,
    criar_notificacao_para_parceiro,
    criar_notificacao_para_utilizador,
    get_notificacoes_utilizador,
    get_notificacoes_nao_lidas_count,
    criar_notificacao
)

# Create tables
Base.metadata.create_all(bind=engine)

app = FastAPI()

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:8080",
        "http://127.0.0.1:8080",
        "http://localhost:8000",
        "http://127.0.0.1:8000",
        "http://localhost:8501",
        "http://127.0.0.1:8501",
        "*"
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"],
)

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

def criar_versao(db: Session, documento: Documento, estado: EstadoDocumento, criado_por: str, comentario: str = ""):
    versao = VersaoDocumento(
        documento_id=documento.id,
        numero_versao=documento.versao_atual,
        dados=documento.dados.copy(),
        estado=estado,
        comentario=comentario,
        criado_por=criado_por
    )
    db.add(versao)
    db.commit()

# -------------------- Authentication --------------------
@app.post("/registar", response_model=Token)
def registar(utilizador: UtilizadorCreate, db: Session = Depends(get_db)):
    user_existente = db.query(Utilizador).filter(Utilizador.username == utilizador.username).first()
    if user_existente:
        raise HTTPException(status_code=400, detail="Username already exists")

    user = Utilizador(
        username=utilizador.username,
        password_hash=hash_password(utilizador.password),
        perfil=utilizador.perfil,
        nome_completo=utilizador.nome_completo
    )
    db.add(user)
    db.commit()
    db.refresh(user)

    access_token = criar_token_acesso(
        data={"sub": user.username},
        expires_delta=timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    )
    return {"access_token": access_token, "token_type": "bearer"}

@app.post("/login", response_model=Token)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    user = db.query(Utilizador).filter(Utilizador.username == form_data.username).first()
    if not user or not verificar_password(form_data.password, user.password_hash):
        raise HTTPException(status_code=400, detail="Invalid username or password")

    access_token = criar_token_acesso(
        data={"sub": user.username},
        expires_delta=timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    )
    return {"access_token": access_token, "token_type": "bearer"}

@app.get("/me", response_model=dict)
def quem_sou_eu(current_user: Utilizador = Depends(get_current_user)):
    return {
        "username": current_user.username,
        "perfil": current_user.perfil.value,
        "nome_completo": current_user.nome_completo
    }

# -------------------- Partners --------------------
@app.get("/parceiros/disponiveis")
def listar_parceiros_disponiveis(
    db: Session = Depends(get_db),
    current_user: Utilizador = Depends(get_current_user)
):
    if current_user.perfil not in [PerfilUtilizador.EMPRESA, PerfilUtilizador.ADMIN]:
        raise HTTPException(status_code=403, detail="Only companies and administrators can list partners")
    
    parceiros = db.query(Utilizador).filter(Utilizador.perfil == PerfilUtilizador.PARCEIRO).all()
    
    return [
        {
            "username": p.username,
            "perfil": p.perfil.value,
            "nome_completo": p.nome_completo
        }
        for p in parceiros
    ]

# -------------------- Documents --------------------
@app.get("/documentos", response_model=List[DocumentoOut])
def listar_documentos(
    estado: Optional[str] = Query(None, description="Filter by status"),
    db: Session = Depends(get_db),
    current_user: Utilizador = Depends(get_current_user)
):
    query = db.query(Documento)
    if estado:
        estado_map = {
            "Draft": EstadoDocumento.RASCUNHO,
            "Submitted": EstadoDocumento.SUBMETIDO,
            "In Review": EstadoDocumento.EM_REVISAO,
            "Changes Requested": EstadoDocumento.ALTERACOES,
            "Approved": EstadoDocumento.APROVADO,
            "Archived": EstadoDocumento.ARQUIVADO
        }
        if estado in estado_map:
            query = query.filter(Documento.estado == estado_map[estado])

    if current_user.perfil == PerfilUtilizador.PARCEIRO:
        query = query.outerjoin(Documento.parceiros).filter(
            or_(
                Utilizador.username == current_user.username,
                Documento.parceiro_id == current_user.username
            )
        )
    elif current_user.perfil == PerfilUtilizador.EMPRESA:
        query = query.filter(Documento.empresa_id == current_user.username)
    
    documentos = query.order_by(Documento.id.desc()).all()
    
    result = []
    for doc in documentos:
        if doc.parceiros:
            parceiros_ids = [p.username for p in doc.parceiros]
        elif doc.parceiro_id:
            parceiros_ids = [doc.parceiro_id]
        else:
            parceiros_ids = []
            
        doc_dict = {
            "id": doc.id,
            "titulo": doc.titulo,
            "empresa_id": doc.empresa_id,
            "estado": doc.estado,
            "versao_atual": doc.versao_atual,
            "dados": doc.dados,
            "created_at": doc.created_at,
            "updated_at": doc.updated_at,
            "parceiro_id": doc.parceiro_id or (parceiros_ids[0] if parceiros_ids else None),
            "parceiros_ids": parceiros_ids
        }
        result.append(doc_dict)
    
    return result

@app.get("/documentos/pesquisar", response_model=List[DocumentoOut])
def pesquisar_documentos(
    q: Optional[str] = Query(None, description="Text to search"),
    estados: Optional[str] = Query(None, description="Filter by status (comma separated)"),
    data_inicio: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    data_fim: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    order_by: Optional[str] = Query("id", description="Field to order by"),
    order_dir: Optional[str] = Query("desc", description="Order direction (asc/desc)"),
    db: Session = Depends(get_db),
    current_user: Utilizador = Depends(get_current_user)
):
    query = db.query(Documento)

    if q:
        q = f"%{q}%"
        try:
            id_int = int(q.replace("%", ""))
            query = query.filter(
                or_(
                    Documento.id == id_int,
                    Documento.titulo.ilike(q),
                    Documento.parceiro_id.ilike(q),
                    Documento.empresa_id.ilike(q)
                )
            )
        except ValueError:
            query = query.filter(
                or_(
                    Documento.titulo.ilike(q),
                    Documento.parceiro_id.ilike(q),
                    Documento.empresa_id.ilike(q)
                )
            )

    if estados:
        estados_lista = [e.strip() for e in estados.split(",") if e.strip()]
        if estados_lista:
            estado_map = {
                "Draft": EstadoDocumento.RASCUNHO,
                "Submitted": EstadoDocumento.SUBMETIDO,
                "In Review": EstadoDocumento.EM_REVISAO,
                "Changes Requested": EstadoDocumento.ALTERACOES,
                "Approved": EstadoDocumento.APROVADO,
                "Archived": EstadoDocumento.ARQUIVADO
            }
            estados_enum = [estado_map[e] for e in estados_lista if e in estado_map]
            if estados_enum:
                query = query.filter(Documento.estado.in_(estados_enum))

    if data_inicio:
        try:
            data_inicio_dt = datetime.strptime(data_inicio, "%Y-%m-%d")
            query = query.filter(Documento.created_at >= data_inicio_dt)
        except ValueError:
            pass

    if data_fim:
        try:
            data_fim_dt = datetime.strptime(data_fim, "%Y-%m-%d")
            data_fim_dt = data_fim_dt.replace(hour=23, minute=59, second=59)
            query = query.filter(Documento.created_at <= data_fim_dt)
        except ValueError:
            pass

    if current_user.perfil == PerfilUtilizador.PARCEIRO:
        query = query.outerjoin(Documento.parceiros).filter(
            or_(
                Utilizador.username == current_user.username,
                Documento.parceiro_id == current_user.username
            )
        )
    elif current_user.perfil == PerfilUtilizador.EMPRESA:
        query = query.filter(Documento.empresa_id == current_user.username)

    order_map = {
        "id": Documento.id,
        "titulo": Documento.titulo,
        "parceiro_id": Documento.parceiro_id,
        "empresa_id": Documento.empresa_id,
        "estado": Documento.estado,
        "created_at": Documento.created_at,
        "updated_at": Documento.updated_at,
        "versao_atual": Documento.versao_atual
    }
    
    campo_ordem = order_map.get(order_by, Documento.id)
    if order_dir.lower() == "asc":
        query = query.order_by(campo_ordem.asc())
    else:
        query = query.order_by(campo_ordem.desc())

    documentos = query.all()
    
    result = []
    for doc in documentos:
        if doc.parceiros:
            parceiros_ids = [p.username for p in doc.parceiros]
        elif doc.parceiro_id:
            parceiros_ids = [doc.parceiro_id]
        else:
            parceiros_ids = []
            
        doc_dict = {
            "id": doc.id,
            "titulo": doc.titulo,
            "empresa_id": doc.empresa_id,
            "estado": doc.estado,
            "versao_atual": doc.versao_atual,
            "dados": doc.dados,
            "created_at": doc.created_at,
            "updated_at": doc.updated_at,
            "parceiro_id": doc.parceiro_id or (parceiros_ids[0] if parceiros_ids else None),
            "parceiros_ids": parceiros_ids
        }
        result.append(doc_dict)
    
    return result

# -------------------- Documents --------------------
@app.post("/documentos/", response_model=DocumentoOut)
def criar_documento(
    doc: DocumentoCreate,
    db: Session = Depends(get_db),
    current_user: Utilizador = Depends(get_current_user)
):
    try:
        if current_user.perfil not in [PerfilUtilizador.EMPRESA, PerfilUtilizador.ADMIN]:
            raise HTTPException(status_code=403, detail="Only companies and administrators can create documents")

        if not doc.parceiros_ids:
            raise HTTPException(status_code=400, detail="At least one partner is required")
        
        parceiros = db.query(Utilizador).filter(
            Utilizador.username.in_(doc.parceiros_ids),
            Utilizador.perfil == PerfilUtilizador.PARCEIRO
        ).all()
        
        if len(parceiros) != len(doc.parceiros_ids):
            raise HTTPException(status_code=400, detail="One or more partners not found or not valid")

        documento = Documento(
            titulo=doc.titulo,
            empresa_id=current_user.username,
            dados=doc.dados,
            imagem_url=doc.imagem_url,
            estado=EstadoDocumento.RASCUNHO,
            versao_atual=1,
            parceiro_id=doc.parceiros_ids[0]
        )
        db.add(documento)
        db.commit()
        db.refresh(documento)

        for parceiro in parceiros:
            documento.parceiros.append(parceiro)
        
        db.commit()

        criar_versao(db, documento, EstadoDocumento.RASCUNHO, criado_por=current_user.username)

        for parceiro in parceiros:
            criar_notificacao_para_utilizador(
                db=db,
                username=parceiro.username,
                titulo="📄 New document created for you",
                mensagem=f"The company {current_user.username} created the document '{doc.titulo}' for you.",
                link=f"/documentos?doc_id={documento.id}",
                icone="📄"
            )

        if documento.parceiros:
            parceiros_ids = [p.username for p in documento.parceiros]
        elif documento.parceiro_id:
            parceiros_ids = [documento.parceiro_id]
        else:
            parceiros_ids = []

        result = {
            "id": documento.id,
            "titulo": documento.titulo,
            "empresa_id": documento.empresa_id,
            "estado": documento.estado,
            "versao_atual": documento.versao_atual,
            "dados": documento.dados,
            "imagem_url": documento.imagem_url,
            "created_at": documento.created_at,
            "updated_at": documento.updated_at,
            "parceiro_id": documento.parceiro_id or (parceiros_ids[0] if parceiros_ids else None),
            "parceiros_ids": parceiros_ids
        }
        
        return result
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error creating document: {e}")
        print(traceback.format_exc())
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error creating document: {str(e)}")

@app.get("/documentos/{doc_id}", response_model=DocumentoOut)
def obter_documento(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: Utilizador = Depends(get_current_user)
):
    doc = db.query(Documento).filter(Documento.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    
    if current_user.perfil == PerfilUtilizador.PARCEIRO:
        is_associated = False
        if doc.parceiros:
            for p in doc.parceiros:
                if p.username == current_user.username:
                    is_associated = True
                    break
        if not is_associated and doc.parceiro_id == current_user.username:
            is_associated = True
        if not is_associated:
            raise HTTPException(status_code=403, detail="Access denied")
    elif current_user.perfil == PerfilUtilizador.EMPRESA:
        if doc.empresa_id != current_user.username:
            raise HTTPException(status_code=403, detail="Access denied")
    
    if doc.parceiros:
        parceiros_ids = [p.username for p in doc.parceiros]
    elif doc.parceiro_id:
        parceiros_ids = [doc.parceiro_id]
    else:
        parceiros_ids = []
    
    result = {
        "id": doc.id,
        "titulo": doc.titulo,
        "empresa_id": doc.empresa_id,
        "estado": doc.estado,
        "versao_atual": doc.versao_atual,
        "dados": doc.dados,
        "imagem_url": doc.imagem_url,
        "created_at": doc.created_at,
        "updated_at": doc.updated_at,
        "parceiro_id": doc.parceiro_id or (parceiros_ids[0] if parceiros_ids else None),
        "parceiros_ids": parceiros_ids
    }
    
    return result

@app.put("/documentos/{doc_id}/editar", response_model=DocumentoOut)
def editar_documento(
    doc_id: int,
    update: DocumentoUpdate,
    db: Session = Depends(get_db),
    current_user: Utilizador = Depends(get_current_user)
):
    try:
        doc = db.query(Documento).filter(Documento.id == doc_id).first()
        if not doc:
            raise HTTPException(status_code=404, detail="Document not found")
        
        if current_user.perfil != PerfilUtilizador.PARCEIRO:
            raise HTTPException(status_code=403, detail="Only partners can edit documents")
        
        is_associated = False
        if doc.parceiros:
            for p in doc.parceiros:
                if p.username == current_user.username:
                    is_associated = True
                    break
        if not is_associated and doc.parceiro_id == current_user.username:
            is_associated = True
        
        if not is_associated:
            raise HTTPException(status_code=403, detail="Only the associated partner can edit")
        
        if doc.estado not in [EstadoDocumento.RASCUNHO, EstadoDocumento.ALTERACOES]:
            raise HTTPException(400, detail=f"Document is in status '{doc.estado}'. Only Draft or Changes Requested documents can be edited.")
        
        doc.dados = update.dados
        # A imagem NÃO é alterada pelo parceiro
        db.commit()
        db.refresh(doc)
        
        if doc.parceiros:
            parceiros_ids = [p.username for p in doc.parceiros]
        elif doc.parceiro_id:
            parceiros_ids = [doc.parceiro_id]
        else:
            parceiros_ids = []
        
        result = {
            "id": doc.id,
            "titulo": doc.titulo,
            "empresa_id": doc.empresa_id,
            "estado": doc.estado,
            "versao_atual": doc.versao_atual,
            "dados": doc.dados,
            "imagem_url": doc.imagem_url,
            "created_at": doc.created_at,
            "updated_at": doc.updated_at,
            "parceiro_id": doc.parceiro_id or (parceiros_ids[0] if parceiros_ids else None),
            "parceiros_ids": parceiros_ids
        }
        
        return result
    except Exception as e:
        print(f"❌ Error editing document: {e}")
        print(traceback.format_exc())
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error editing: {str(e)}")

@app.post("/documentos/{doc_id}/submeter", response_model=DocumentoOut)
def submeter_documento(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: Utilizador = Depends(get_current_user)
):
    try:
        doc = db.query(Documento).filter(Documento.id == doc_id).first()
        if not doc:
            raise HTTPException(status_code=404, detail="Document not found")
        
        if current_user.perfil != PerfilUtilizador.PARCEIRO:
            raise HTTPException(status_code=403, detail="Only partners can submit documents")
        
        is_associated = False
        if doc.parceiros:
            for p in doc.parceiros:
                if p.username == current_user.username:
                    is_associated = True
                    break
        if not is_associated and doc.parceiro_id == current_user.username:
            is_associated = True
        
        if not is_associated:
            raise HTTPException(status_code=403, detail="Only the associated partner can submit")
        
        if doc.estado not in [EstadoDocumento.RASCUNHO, EstadoDocumento.ALTERACOES]:
            raise HTTPException(400, detail=f"Can only submit from Draft or Changes Requested status. Current status: {doc.estado.value}")
        
        criar_versao(db, doc, EstadoDocumento.SUBMETIDO, criado_por=current_user.username, comentario="Submission for validation")
        
        doc.versao_atual += 1
        doc.estado = EstadoDocumento.SUBMETIDO
        
        db.commit()
        db.refresh(doc)
        
        criar_notificacao_para_empresa(
            db=db,
            documento=doc,
            titulo="📤 Document submitted",
            mensagem=f"The document '{doc.titulo}' was submitted by {current_user.username} for validation.",
            icone="📤"
        )
        
        if doc.parceiros:
            parceiros_ids = [p.username for p in doc.parceiros]
        elif doc.parceiro_id:
            parceiros_ids = [doc.parceiro_id]
        else:
            parceiros_ids = []
        
        result = {
            "id": doc.id,
            "titulo": doc.titulo,
            "empresa_id": doc.empresa_id,
            "estado": doc.estado,
            "versao_atual": doc.versao_atual,
            "dados": doc.dados,
            "created_at": doc.created_at,
            "updated_at": doc.updated_at,
            "parceiro_id": doc.parceiro_id or (parceiros_ids[0] if parceiros_ids else None),
            "parceiros_ids": parceiros_ids
        }
        
        return result
    except Exception as e:
        print(f"❌ Error submitting document: {e}")
        print(traceback.format_exc())
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error submitting: {str(e)}")

@app.post("/documentos/{doc_id}/iniciar-revisao", response_model=DocumentoOut)
def iniciar_revisao(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: Utilizador = Depends(get_current_user)
):
    doc = db.query(Documento).filter(Documento.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    
    if current_user.perfil not in [PerfilUtilizador.EMPRESA, PerfilUtilizador.ADMIN]:
        raise HTTPException(status_code=403, detail="Only companies and administrators can start review")
    
    if current_user.perfil == PerfilUtilizador.EMPRESA and doc.empresa_id != current_user.username:
        raise HTTPException(status_code=403, detail="Only the company that created the document can start review")
    
    if doc.estado != EstadoDocumento.SUBMETIDO:
        raise HTTPException(400, detail="Document is not submitted")
    
    doc.estado = EstadoDocumento.EM_REVISAO
    criar_versao(db, doc, EstadoDocumento.EM_REVISAO, criado_por=current_user.username, comentario="Review started")
    db.commit()
    db.refresh(doc)
    
    criar_notificacao_para_parceiro(
        db=db,
        documento=doc,
        titulo="🔍 Review started",
        mensagem=f"The document '{doc.titulo}' is under review by the company.",
        icone="🔍"
    )

    return doc

@app.post("/documentos/{doc_id}/pedir-alteracoes", response_model=DocumentoOut)
def pedir_alteracoes(
    doc_id: int,
    motivo: MudancaEstado,
    db: Session = Depends(get_db),
    current_user: Utilizador = Depends(get_current_user)
):
    doc = db.query(Documento).filter(Documento.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    
    if current_user.perfil not in [PerfilUtilizador.EMPRESA, PerfilUtilizador.ADMIN]:
        raise HTTPException(status_code=403, detail="Only companies and administrators can request changes")
    
    if current_user.perfil == PerfilUtilizador.EMPRESA and doc.empresa_id != current_user.username:
        raise HTTPException(status_code=403, detail="Only the company that created the document can request changes")
    
    if doc.estado != EstadoDocumento.EM_REVISAO:
        raise HTTPException(400, detail="Can only request changes during review")
    
    doc.estado = EstadoDocumento.ALTERACOES
    criar_versao(db, doc, EstadoDocumento.ALTERACOES, criado_por=current_user.username, comentario=motivo.comentario or "Changes requested")
    db.commit()
    db.refresh(doc)
    
    criar_notificacao_para_parceiro(
        db=db,
        documento=doc,
        titulo="🔄 Changes requested",
        mensagem=f"The company requested changes for '{doc.titulo}': {motivo.comentario or 'See details'}",
        icone="🔄"
    )
    
    return doc

@app.post("/documentos/{doc_id}/editar-novamente", response_model=DocumentoOut)
def editar_novamente(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: Utilizador = Depends(get_current_user)
):
    doc = db.query(Documento).filter(Documento.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    
    if current_user.perfil != PerfilUtilizador.PARCEIRO:
        raise HTTPException(status_code=403, detail="Only the associated partner can edit again")
    
    is_associated = False
    if doc.parceiros:
        for p in doc.parceiros:
            if p.username == current_user.username:
                is_associated = True
                break
    if not is_associated and doc.parceiro_id == current_user.username:
        is_associated = True
    
    if not is_associated:
        raise HTTPException(status_code=403, detail="Only the associated partner can edit again")
    
    if doc.estado != EstadoDocumento.ALTERACOES:
        raise HTTPException(400, detail="Action only allowed in 'Changes Requested' status")
    
    doc.estado = EstadoDocumento.RASCUNHO
    doc.versao_atual += 1
    criar_versao(db, doc, EstadoDocumento.RASCUNHO, criado_por=current_user.username, comentario="Started corrections after change request")
    db.commit()
    db.refresh(doc)
    return doc

@app.post("/documentos/{doc_id}/aprovar", response_model=DocumentoOut)
def aprovar_documento(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: Utilizador = Depends(get_current_user)
):
    doc = db.query(Documento).filter(Documento.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    
    if current_user.perfil not in [PerfilUtilizador.EMPRESA, PerfilUtilizador.ADMIN]:
        raise HTTPException(status_code=403, detail="Only companies and administrators can approve")
    
    if current_user.perfil == PerfilUtilizador.EMPRESA and doc.empresa_id != current_user.username:
        raise HTTPException(status_code=403, detail="Only the company that created the document can approve")
    
    if doc.estado != EstadoDocumento.EM_REVISAO:
        raise HTTPException(400, detail="Can only approve during review")
    
    doc.estado = EstadoDocumento.APROVADO
    criar_versao(db, doc, EstadoDocumento.APROVADO, criado_por=current_user.username, comentario="Document approved")
    db.commit()
    db.refresh(doc)
    
    criar_notificacao_para_parceiro(
        db=db,
        documento=doc,
        titulo="✅ Document approved",
        mensagem=f"The document '{doc.titulo}' was approved by {current_user.username}.",
        icone="✅"
    )

    return doc

@app.post("/documentos/{doc_id}/reabrir", response_model=DocumentoOut)
def reabrir_documento(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: Utilizador = Depends(get_current_user)
):
    doc = db.query(Documento).filter(Documento.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    
    if current_user.perfil not in [PerfilUtilizador.EMPRESA, PerfilUtilizador.ADMIN]:
        raise HTTPException(status_code=403, detail="Only companies and administrators can reopen")
    
    if current_user.perfil == PerfilUtilizador.EMPRESA and doc.empresa_id != current_user.username:
        raise HTTPException(status_code=403, detail="Only the company that created the document can reopen")
    
    if doc.estado != EstadoDocumento.APROVADO:
        raise HTTPException(400, detail="Can only reopen an approved document")
    
    doc.estado = EstadoDocumento.RASCUNHO
    doc.versao_atual += 1
    criar_versao(db, doc, EstadoDocumento.RASCUNHO, criado_por=current_user.username, comentario="Document reopened for new editing")
    db.commit()
    db.refresh(doc)
    return doc

@app.post("/documentos/{doc_id}/arquivar", response_model=DocumentoOut)
def arquivar_documento(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: Utilizador = Depends(get_current_user)
):
    doc = db.query(Documento).filter(Documento.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    
    if current_user.perfil not in [PerfilUtilizador.EMPRESA, PerfilUtilizador.ADMIN]:
        raise HTTPException(status_code=403, detail="Only companies and administrators can archive documents")
    
    if current_user.perfil == PerfilUtilizador.EMPRESA and doc.empresa_id != current_user.username:
        raise HTTPException(status_code=403, detail="Only the company that created the document can archive")
    
    if doc.estado not in [EstadoDocumento.RASCUNHO, EstadoDocumento.APROVADO]:
        raise HTTPException(400, detail="Can only archive documents in Draft or Approved status")
    
    doc.estado = EstadoDocumento.ARQUIVADO
    criar_versao(db, doc, EstadoDocumento.ARQUIVADO, criado_por=current_user.username, comentario="Document archived")
    db.commit()
    db.refresh(doc)
    
    criar_notificacao_para_parceiro(
        db=db,
        documento=doc,
        titulo="📁 Document archived",
        mensagem=f"The document '{doc.titulo}' was archived by {current_user.username}.",
        icone="📁"
    )

    return doc

@app.get("/documentos/{doc_id}/versoes", response_model=List[VersaoOut])
def listar_versoes(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: Utilizador = Depends(get_current_user)
):
    doc = db.query(Documento).filter(Documento.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    
    if current_user.perfil == PerfilUtilizador.PARCEIRO:
        is_associated = False
        if doc.parceiros:
            for p in doc.parceiros:
                if p.username == current_user.username:
                    is_associated = True
                    break
        if not is_associated and doc.parceiro_id == current_user.username:
            is_associated = True
        if not is_associated:
            raise HTTPException(status_code=403, detail="Access denied")
    elif current_user.perfil == PerfilUtilizador.EMPRESA:
        if doc.empresa_id != current_user.username:
            raise HTTPException(status_code=403, detail="Access denied")
    
    versoes = db.query(VersaoDocumento).filter(VersaoDocumento.documento_id == doc_id).order_by(VersaoDocumento.numero_versao).all()
    return versoes

# -------------------- Comments --------------------
@app.post("/documentos/{doc_id}/comentarios", response_model=ComentarioOut)
def adicionar_comentario(
    doc_id: int,
    comentario: ComentarioCreate,
    db: Session = Depends(get_db),
    current_user: Utilizador = Depends(get_current_user)
):
    doc = db.query(Documento).filter(Documento.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    
    if current_user.perfil == PerfilUtilizador.PARCEIRO:
        is_associated = False
        if doc.parceiros:
            for p in doc.parceiros:
                if p.username == current_user.username:
                    is_associated = True
                    break
        if not is_associated and doc.parceiro_id == current_user.username:
            is_associated = True
        if not is_associated:
            raise HTTPException(status_code=403, detail="Access denied")
    elif current_user.perfil == PerfilUtilizador.EMPRESA:
        if doc.empresa_id != current_user.username:
            raise HTTPException(status_code=403, detail="Access denied")
    
    novo_comentario = ComentarioDocumento(
        documento_id=doc_id,
        username=current_user.username,
        mensagem=comentario.mensagem
    )
    db.add(novo_comentario)
    db.commit()
    db.refresh(novo_comentario)
    
    usuarios_notificar = []
    
    if current_user.username != doc.empresa_id:
        usuarios_notificar.append(doc.empresa_id)
    
    if doc.parceiros:
        for p in doc.parceiros:
            if p.username != current_user.username and p.username not in usuarios_notificar:
                usuarios_notificar.append(p.username)
    elif doc.parceiro_id and doc.parceiro_id != current_user.username:
        if doc.parceiro_id not in usuarios_notificar:
            usuarios_notificar.append(doc.parceiro_id)
    
    if current_user.perfil != PerfilUtilizador.ADMIN:
        admins = db.query(Utilizador).filter(Utilizador.perfil == PerfilUtilizador.ADMIN).all()
        for admin in admins:
            if admin.username != current_user.username and admin.username not in usuarios_notificar:
                usuarios_notificar.append(admin.username)
    
    for username in usuarios_notificar:
        criar_notificacao(
            db=db,
            username=username,
            titulo=f"💬 New comment on document",
            mensagem=f"{current_user.username} commented on '{doc.titulo}': {comentario.mensagem[:100]}{'...' if len(comentario.mensagem) > 100 else ''}",
            link=f"/documentos?doc_id={doc_id}",
            icone="💬"
        )
    
    return novo_comentario.to_dict()

@app.get("/documentos/{doc_id}/comentarios", response_model=List[ComentarioOut])
def listar_comentarios(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: Utilizador = Depends(get_current_user)
):
    doc = db.query(Documento).filter(Documento.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    
    if current_user.perfil == PerfilUtilizador.PARCEIRO:
        is_associated = False
        if doc.parceiros:
            for p in doc.parceiros:
                if p.username == current_user.username:
                    is_associated = True
                    break
        if not is_associated and doc.parceiro_id == current_user.username:
            is_associated = True
        if not is_associated:
            raise HTTPException(status_code=403, detail="Access denied")
    elif current_user.perfil == PerfilUtilizador.EMPRESA:
        if doc.empresa_id != current_user.username:
            raise HTTPException(status_code=403, detail="Access denied")
    
    comentarios = db.query(ComentarioDocumento).filter(
        ComentarioDocumento.documento_id == doc_id
    ).order_by(ComentarioDocumento.created_at.asc()).all()
    
    return [c.to_dict() for c in comentarios]


# -------------------- Admin: user management --------------------
@app.get("/admin/usuarios", response_model=List[dict])
def listar_utilizadores(
    db: Session = Depends(get_db),
    current_user: Utilizador = Depends(get_current_user)
):
    if current_user.perfil != PerfilUtilizador.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    users = db.query(Utilizador).all()
    return [
        {
            "username": u.username,
            "perfil": u.perfil.value,
            "nome_completo": u.nome_completo,
            "created_at": u.created_at
        }
        for u in users
    ]

@app.delete("/admin/usuarios/{username}")
def eliminar_utilizador(
    username: str,
    db: Session = Depends(get_db),
    current_user: Utilizador = Depends(get_current_user)
):
    if current_user.perfil != PerfilUtilizador.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    if username == current_user.username:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")
    user = db.query(Utilizador).filter(Utilizador.username == username).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    db.delete(user)
    db.commit()
    return {"ok": True}

@app.put("/admin/usuarios/{username}/password")
def alterar_password(
    username: str,
    dados: PasswordUpdate,
    db: Session = Depends(get_db),
    current_user: Utilizador = Depends(get_current_user)
):
    if current_user.perfil != PerfilUtilizador.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    user = db.query(Utilizador).filter(Utilizador.username == username).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    user.password_hash = hash_password(dados.nova_password)
    db.commit()
    return {"ok": True}

# -------------------- Admin: delete document --------------------
@app.delete("/admin/documentos/{doc_id}")
def eliminar_documento(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: Utilizador = Depends(get_current_user)
):
    if current_user.perfil != PerfilUtilizador.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    doc = db.query(Documento).filter(Documento.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    
    versoes = db.query(VersaoDocumento).filter(VersaoDocumento.documento_id == doc_id).all()
    for v in versoes:
        db.delete(v)
    
    notificacoes = db.query(Notificacao).filter(Notificacao.link.like(f"%doc_id={doc_id}%")).all()
    for n in notificacoes:
        db.delete(n)
    
    comentarios = db.query(ComentarioDocumento).filter(ComentarioDocumento.documento_id == doc_id).all()
    for c in comentarios:
        db.delete(c)
    
    db.delete(doc)
    db.commit()
    
    return {"ok": True, "message": f"Document {doc_id} deleted successfully"}

# -------------------- Dashboard --------------------
@app.get("/dashboard/kpis")
def dashboard_kpis(
    db: Session = Depends(get_db),
    current_user: Utilizador = Depends(get_current_user)
):
    return get_dashboard_kpis(db, current_user.username, current_user.perfil)

@app.get("/dashboard/top-parceiros")
def dashboard_top_parceiros(
    limit: int = 10,
    db: Session = Depends(get_db),
    current_user: Utilizador = Depends(get_current_user)
):
    if current_user.perfil == PerfilUtilizador.PARCEIRO:
        raise HTTPException(status_code=403, detail="Only companies/admins can view top partners")
    return get_top_parceiros(db, limit)

@app.get("/dashboard/documentos-recentes")
def dashboard_documentos_recentes(
    limit: int = Query(None, description="Number of documents to return (optional)"),
    db: Session = Depends(get_db),
    current_user: Utilizador = Depends(get_current_user)
):
    return get_documentos_recentes(db, current_user.username, current_user.perfil, limit)

# -------------------- Notifications --------------------
@app.get("/notificacoes")
def listar_notificacoes(
    limit: int = 50,
    db: Session = Depends(get_db),
    current_user: Utilizador = Depends(get_current_user)
):
    return get_notificacoes_utilizador(db, current_user.username, limit)

@app.get("/notificacoes/nao-lidas")
def contar_notificacoes_nao_lidas(
    db: Session = Depends(get_db),
    current_user: Utilizador = Depends(get_current_user)
):
    count = get_notificacoes_nao_lidas_count(db, current_user.username)
    return {"count": count}

@app.put("/notificacoes/{notificacao_id}/ler")
def marcar_notificacao_lida(
    notificacao_id: int,
    db: Session = Depends(get_db),
    current_user: Utilizador = Depends(get_current_user)
):
    try:
        notificacao = db.query(Notificacao).filter(
            Notificacao.id == notificacao_id,
            Notificacao.username == current_user.username
        ).first()
        
        if not notificacao:
            raise HTTPException(status_code=404, detail="Notification not found")
        
        notificacao.lida = True
        db.commit()
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error marking notification as read: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/notificacoes/ler-todas")
def marcar_todas_notificacoes_lidas(
    db: Session = Depends(get_db),
    current_user: Utilizador = Depends(get_current_user)
):
    try:
        count = db.query(Notificacao).filter(
            Notificacao.username == current_user.username,
            Notificacao.lida == False
        ).update({"lida": True})
        db.commit()
        return {"ok": True, "count": count}
    except Exception as e:
        print(f"❌ Error marking all notifications as read: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# -------------------- Export Excel --------------------
@app.get("/documentos/{doc_id}/exportar-excel")
def exportar_versoes_excel(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: Utilizador = Depends(get_current_user)
):
    try:
        doc = db.query(Documento).filter(Documento.id == doc_id).first()
        if not doc:
            raise HTTPException(status_code=404, detail="Document not found")
        
        if current_user.perfil == PerfilUtilizador.PARCEIRO:
            is_associated = False
            if doc.parceiros:
                for p in doc.parceiros:
                    if p.username == current_user.username:
                        is_associated = True
                        break
            if not is_associated and doc.parceiro_id == current_user.username:
                is_associated = True
            if not is_associated:
                raise HTTPException(status_code=403, detail="Access denied")
        elif current_user.perfil == PerfilUtilizador.EMPRESA:
            if doc.empresa_id != current_user.username:
                raise HTTPException(status_code=403, detail="Access denied")

        dados = doc.dados
        wb = openpyxl.Workbook()
        
        processos = get_processos_from_data(dados)

        # LCA Sheet
        ws_lca = wb.active
        ws_lca.title = "LCA"
        row = 1

        ws_lca.cell(row=row, column=1, value="INPUTS").font = Font(bold=True, size=12)
        row += 1
        cab_inputs = ["Process", "Material", "QTY", "Unit", "Material Description", "CAS/Comments", "Distance (km)", "Country", "Data Source"]
        for col, header in enumerate(cab_inputs, start=1):
            ws_lca.cell(row=row, column=col, value=header).font = Font(bold=True)
        row += 1
        for proc in processos:
            for item in dados.get("lca", {}).get("inputs", {}).get(proc, []):
                ws_lca.cell(row=row, column=1, value=proc)
                ws_lca.cell(row=row, column=2, value=item.get("material", ""))
                ws_lca.cell(row=row, column=3, value=item.get("qty", ""))
                ws_lca.cell(row=row, column=4, value=item.get("unit", ""))
                ws_lca.cell(row=row, column=5, value=item.get("description", ""))
                ws_lca.cell(row=row, column=6, value=item.get("cas", ""))
                ws_lca.cell(row=row, column=7, value=item.get("distance", ""))
                ws_lca.cell(row=row, column=8, value=item.get("country", ""))
                ws_lca.cell(row=row, column=9, value=item.get("datasource", ""))
                row += 1
        row += 1

        ws_lca.cell(row=row, column=1, value="PROCESSES").font = Font(bold=True, size=12)
        row += 1
        cab_proc = ["Process", "Type", "QTY", "Unit", "Description", "Comments", "Data Source"]
        for col, header in enumerate(cab_proc, start=1):
            ws_lca.cell(row=row, column=col, value=header).font = Font(bold=True)
        row += 1
        for proc in processos:
            for item in dados.get("lca", {}).get("processes", {}).get(proc, []):
                ws_lca.cell(row=row, column=1, value=proc)
                ws_lca.cell(row=row, column=2, value=item.get("tipo", ""))
                ws_lca.cell(row=row, column=3, value=item.get("qty", ""))
                ws_lca.cell(row=row, column=4, value=item.get("unit", ""))
                ws_lca.cell(row=row, column=5, value=item.get("description", ""))
                ws_lca.cell(row=row, column=6, value=item.get("comments", ""))
                ws_lca.cell(row=row, column=7, value=item.get("datasource", ""))
                row += 1
        row += 1

        ws_lca.cell(row=row, column=1, value="OUTPUTS").font = Font(bold=True, size=12)
        row += 1
        cab_out = ["Process", "Step", "Type", "Sub-type", "QTY", "Unit", "Material description", "Comments", "Data Source"]
        for col, header in enumerate(cab_out, start=1):
            ws_lca.cell(row=row, column=col, value=header).font = Font(bold=True)
        row += 1
        for proc in processos:
            for item in dados.get("lca", {}).get("outputs", {}).get(proc, []):
                ws_lca.cell(row=row, column=1, value=proc)
                ws_lca.cell(row=row, column=2, value=item.get("etapa", ""))
                ws_lca.cell(row=row, column=3, value=item.get("tipo", ""))
                ws_lca.cell(row=row, column=4, value=item.get("sub_tipo", ""))
                ws_lca.cell(row=row, column=5, value=item.get("qty", ""))
                ws_lca.cell(row=row, column=6, value=item.get("unit", ""))
                ws_lca.cell(row=row, column=7, value=item.get("description", ""))
                ws_lca.cell(row=row, column=8, value=item.get("comments", ""))
                ws_lca.cell(row=row, column=9, value=item.get("datasource", ""))
                row += 1

        for col in ws_lca.columns:
            max_length = 0
            for cell in col:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            col_letter = get_column_letter(cell.column)
            ws_lca.column_dimensions[col_letter].width = min(max_length + 2, 40)

        # LCC Sheet
        ws_lcc = wb.create_sheet("LCC")
        row = 1

        ws_lcc.cell(row=row, column=1, value="COST BREAKDOWN MATERIAL").font = Font(bold=True, size=12)
        row += 1
        cab_mat = ["Process", "Material", "Price €", "Qty", "Unit", "Material Description", "Comments", "Distance (km)", "Country", "Data Source"]
        for col, header in enumerate(cab_mat, start=1):
            ws_lcc.cell(row=row, column=col, value=header).font = Font(bold=True)
        row += 1
        for proc in processos:
            for item in dados.get("lcc", {}).get("materials", {}).get(proc, []):
                ws_lcc.cell(row=row, column=1, value=proc)
                ws_lcc.cell(row=row, column=2, value=item.get("material", ""))
                ws_lcc.cell(row=row, column=3, value=item.get("price", ""))
                ws_lcc.cell(row=row, column=4, value=item.get("qty", ""))
                ws_lcc.cell(row=row, column=5, value=item.get("unit", ""))
                ws_lcc.cell(row=row, column=6, value=item.get("description", ""))
                ws_lcc.cell(row=row, column=7, value=item.get("comments", ""))
                ws_lcc.cell(row=row, column=8, value=item.get("distance", ""))
                ws_lcc.cell(row=row, column=9, value=item.get("country", ""))
                ws_lcc.cell(row=row, column=10, value=item.get("datasource", ""))
                row += 1
        row += 1

        ws_lcc.cell(row=row, column=1, value="EQUIPMENT").font = Font(bold=True, size=12)
        row += 1
        cab_eq = ["Process", "Equipment", "Process", "Unit Cost (€)", "Lifespan (Years)", "Maintenance €/Year", "Industrial Equivalent", "Comments", "Data Source"]
        for col, header in enumerate(cab_eq, start=1):
            ws_lcc.cell(row=row, column=col, value=header).font = Font(bold=True)
        row += 1
        for proc in processos:
            for item in dados.get("lcc", {}).get("equipment", {}).get(proc, []):
                ws_lcc.cell(row=row, column=1, value=proc)
                ws_lcc.cell(row=row, column=2, value=item.get("equipment", ""))
                ws_lcc.cell(row=row, column=3, value=item.get("process", ""))
                ws_lcc.cell(row=row, column=4, value=item.get("unit_cost", ""))
                ws_lcc.cell(row=row, column=5, value=item.get("lifespan", ""))
                ws_lcc.cell(row=row, column=6, value=item.get("maintenance", ""))
                ws_lcc.cell(row=row, column=7, value=item.get("industrial_equiv", ""))
                ws_lcc.cell(row=row, column=8, value=item.get("comments", ""))
                ws_lcc.cell(row=row, column=9, value=item.get("datasource", ""))
                row += 1
        row += 1

        ws_lcc.cell(row=row, column=1, value="LABOUR").font = Font(bold=True, size=12)
        row += 1
        cab_lab = ["Process", "Name of the process", "Total Labour - Number", "Total Labour - Cost €",
                   "Number - High Skilled", "Number - Moderated skilled", "Number - Unskilled",
                   "Rate - High Skilled (€/h)", "Rate - Moderated skilled (€/h)", "Rate - Unskilled (€/h)",
                   "Comments", "Data Source"]
        for col, header in enumerate(cab_lab, start=1):
            ws_lcc.cell(row=row, column=col, value=header).font = Font(bold=True)
        row += 1
        for proc in processos:
            for item in dados.get("lcc", {}).get("labour", {}).get(proc, []):
                ws_lcc.cell(row=row, column=1, value=proc)
                ws_lcc.cell(row=row, column=2, value=item.get("process", ""))
                ws_lcc.cell(row=row, column=3, value=item.get("total_number", ""))
                ws_lcc.cell(row=row, column=4, value=item.get("total_cost", ""))
                ws_lcc.cell(row=row, column=5, value=item.get("high_skilled", ""))
                ws_lcc.cell(row=row, column=6, value=item.get("moderate_skilled", ""))
                ws_lcc.cell(row=row, column=7, value=item.get("unskilled", ""))
                ws_lcc.cell(row=row, column=8, value=item.get("high_rate", ""))
                ws_lcc.cell(row=row, column=9, value=item.get("moderate_rate", ""))
                ws_lcc.cell(row=row, column=10, value=item.get("unskilled_rate", ""))
                ws_lcc.cell(row=row, column=11, value=item.get("comments", ""))
                ws_lcc.cell(row=row, column=12, value=item.get("datasource", ""))
                row += 1
        row += 1

        ws_lcc.cell(row=row, column=1, value="OUTPUTS").font = Font(bold=True, size=12)
        row += 1
        cab_out_lcc = ["Process", "Material", "Market Price €", "Quantity", "Unit", "Amount of product produced", "Comments", "Data Source"]
        for col, header in enumerate(cab_out_lcc, start=1):
            ws_lcc.cell(row=row, column=col, value=header).font = Font(bold=True)
        row += 1
        for proc in processos:
            for item in dados.get("lcc", {}).get("outputs", {}).get(proc, []):
                ws_lcc.cell(row=row, column=1, value=proc)
                ws_lcc.cell(row=row, column=2, value=item.get("material", ""))
                ws_lcc.cell(row=row, column=3, value=item.get("market_price", ""))
                ws_lcc.cell(row=row, column=4, value=item.get("quantity", ""))
                ws_lcc.cell(row=row, column=5, value=item.get("unit", ""))
                ws_lcc.cell(row=row, column=6, value=item.get("amount_produced", ""))
                ws_lcc.cell(row=row, column=7, value=item.get("comments", ""))
                ws_lcc.cell(row=row, column=8, value=item.get("datasource", ""))
                row += 1

        for col in ws_lcc.columns:
            max_length = 0
            for cell in col:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            col_letter = get_column_letter(cell.column)
            ws_lcc.column_dimensions[col_letter].width = min(max_length + 2, 40)

        # History
        ws_hist = wb.create_sheet("History")
        versoes = db.query(VersaoDocumento).filter(VersaoDocumento.documento_id == doc_id).order_by(VersaoDocumento.numero_versao).all()
        cab_hist = ["Version", "Status", "Created by", "Date", "Comment"]
        for col, header in enumerate(cab_hist, start=1):
            ws_hist.cell(row=1, column=col, value=header).font = Font(bold=True)
        for idx, v in enumerate(versoes, start=2):
            ws_hist.cell(row=idx, column=1, value=v.numero_versao)
            ws_hist.cell(row=idx, column=2, value=v.estado.value if v.estado else "")
            ws_hist.cell(row=idx, column=3, value=v.criado_por or "")
            ws_hist.cell(row=idx, column=4, value=v.created_at.strftime("%Y-%m-%d %H:%M:%S") if v.created_at else "")
            ws_hist.cell(row=idx, column=5, value=v.comentario or "")
        for col in ws_hist.columns:
            max_length = 0
            for cell in col:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            col_letter = get_column_letter(cell.column)
            ws_hist.column_dimensions[col_letter].width = min(max_length + 2, 40)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"{doc.titulo}.xlsx"
        filename = "".join(c for c in filename if c.isalnum() or c in " ._-")
        
        headers = {"Content-Disposition": f"attachment; filename={filename}"}
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

    except Exception as e:
        return JSONResponse(status_code=500, content={"detail": f"Error generating Excel: {str(e)}"})